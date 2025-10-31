"""
Módulo para exportação de relatórios em Excel
Gera arquivos Excel formatados com múltiplas abas e gráficos
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from io import BytesIO
from datetime import datetime
from typing import Dict, List
from analytics import AgendorAnalytics


def apply_header_style(ws, max_col):
    """Aplica estilo ao cabeçalho (primeira linha)"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_borders(ws, max_row, max_col):
    """Aplica bordas a todas as células"""
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def auto_adjust_columns(ws, max_col):
    """Ajusta largura das colunas automaticamente"""
    for col in range(1, max_col + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col).column_letter
        
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_kpi_sheet(wb, analytics: AgendorAnalytics):
    """Cria aba de KPIs principais"""
    ws = wb.create_sheet("📊 KPIs Principais")
    
    # Calcular métricas
    revenue = analytics.calculate_revenue_forecast()
    win_loss = analytics.calculate_win_loss_rate()
    time_to_close = analytics.calculate_average_time_to_close()
    growth = analytics.calculate_growth_trend()
    
    # Criar DataFrame
    data = {
        'Métrica': [
            'Receita Confirmada (Ganhos)',
            'Receita em Andamento',
            'Previsão de Receita',
            'Taxa de Vitória',
            'Taxa de Perda',
            'Negócios Ganhos',
            'Negócios Perdidos',
            'Tempo Médio para Ganhar (dias)',
            'Tempo Médio para Perder (dias)',
            'Crescimento Receita (30 dias)',
            'Receita Últimos 30 dias'
        ],
        'Valor': [
            f"R$ {revenue.get('receita_confirmada', 0):,.2f}",
            f"R$ {revenue.get('receita_em_andamento', 0):,.2f}",
            f"R$ {revenue.get('previsao_receita', 0):,.2f}",
            f"{win_loss.get('taxa_vitoria', 0):.1f}%",
            f"{win_loss.get('taxa_perda', 0):.1f}%",
            f"{win_loss.get('ganhos', 0)}",
            f"{win_loss.get('perdidos', 0)}",
            f"{time_to_close.get('tempo_medio_ganhos', 0):.0f}",
            f"{time_to_close.get('tempo_medio_perdas', 0):.0f}",
            f"{growth.get('crescimento_percentual', 0):.1f}%",
            f"R$ {growth.get('receita_ultimos_30_dias', 0):,.2f}"
        ]
    }
    
    df = pd.DataFrame(data)
    
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:  # Não centralizar cabeçalho
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Aplicar estilos
    apply_header_style(ws, 2)
    apply_borders(ws, len(df) + 1, 2)
    auto_adjust_columns(ws, 2)
    
    # Destacar valores importantes em verde/vermelho
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Taxa de vitória (verde se > 50%)
    if win_loss.get('taxa_vitoria', 0) > 50:
        ws['B5'].fill = green_fill
    
    # Crescimento (verde se positivo)
    if growth.get('crescimento_percentual', 0) > 0:
        ws['B11'].fill = green_fill
    elif growth.get('crescimento_percentual', 0) < 0:
        ws['B11'].fill = red_fill


def create_sellers_sheet(wb, analytics: AgendorAnalytics):
    """Cria aba de performance por vendedor com gráfico"""
    ws = wb.create_sheet("👥 Vendedores")
    
    sellers_data = analytics.calculate_seller_performance()
    
    if not sellers_data:
        ws['A1'] = "Sem dados de vendedores"
        return
    
    # Criar DataFrame
    data = []
    for seller in sellers_data:
        data.append({
            'Vendedor': seller['vendedor'],
            'Receita Total': seller['receita_total'],
            'Negócios Ganhos': seller['negocios_ganhos'],
            'Negócios Perdidos': seller['negocios_perdidos'],
            'Taxa de Vitória (%)': seller['taxa_vitoria'],
            'Ticket Médio': seller['ticket_medio'],
            'Tempo Médio (dias)': seller['tempo_medio_fechamento']
        })
    
    df = pd.DataFrame(data)
    
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Aplicar estilos
    apply_header_style(ws, 7)
    apply_borders(ws, len(df) + 1, 7)
    auto_adjust_columns(ws, 7)
    
    # Formatar valores monetários
    for row in range(2, len(df) + 2):
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'F{row}'].number_format = 'R$ #,##0.00'
    
    # Criar gráfico de barras - Receita por Vendedor
    chart = BarChart()
    chart.title = "Receita por Vendedor"
    chart.style = 10
    chart.y_axis.title = 'Receita (R$)'
    chart.x_axis.title = 'Vendedor'
    
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 20
    
    ws.add_chart(chart, "I2")


def create_top_customers_sheet(wb, analytics: AgendorAnalytics):
    """Cria aba de Top Clientes"""
    ws = wb.create_sheet("🏆 Top Clientes")
    
    top_customers = analytics.calculate_top_customers(limit=20)
    
    if not top_customers:
        ws['A1'] = "Sem dados de clientes"
        return
    
    # Criar DataFrame
    data = []
    for customer in top_customers:
        data.append({
            'Cliente': customer['cliente'],
            'Receita Total': customer['receita_total'],
            'Quantidade de Negócios': customer['quantidade_negocios'],
            'Ticket Médio': customer['ticket_medio']
        })
    
    df = pd.DataFrame(data)
    
    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Aplicar estilos
    apply_header_style(ws, 4)
    apply_borders(ws, len(df) + 1, 4)
    auto_adjust_columns(ws, 4)
    
    # Formatar valores monetários
    for row in range(2, len(df) + 2):
        ws[f'B{row}'].number_format = 'R$ #,##0.00'
        ws[f'D{row}'].number_format = 'R$ #,##0.00'
    
    # Criar gráfico de pizza - Top 10
    pie = PieChart()
    pie.title = "Top 10 Clientes - Distribuição de Receita"
    
    labels = Reference(ws, min_col=1, min_row=2, max_row=min(11, len(df) + 1))
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=min(11, len(df) + 1))
    
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 12
    pie.width = 16
    
    ws.add_chart(pie, "F2")


def create_funnel_sheet(wb, analytics: AgendorAnalytics):
    """Cria aba de Funil de Conversão"""
    ws = wb.create_sheet("🎯 Funil")
    
    conversion_data = analytics.calculate_conversion_rates()
    
    if conversion_data.empty:
        ws['A1'] = "Sem dados de funil"
        return
    
    # Agrupar por funil
    for funnel_name in conversion_data['funil'].unique():
        funnel_df = conversion_data[conversion_data['funil'] == funnel_name].sort_values('ordem')
        
        # Escrever nome do funil
        current_row = ws.max_row + 2
        ws[f'A{current_row}'] = f"Funil: {funnel_name}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        
        # Cabeçalho
        current_row += 1
        headers = ['Etapa', 'Quantidade', 'Taxa de Conversão (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Dados
        for _, row in funnel_df.iterrows():
            current_row += 1
            ws[f'A{current_row}'] = row['etapa']
            ws[f'B{current_row}'] = row['quantidade']
            ws[f'C{current_row}'] = row['taxa_conversao']
            ws[f'C{current_row}'].number_format = '0.00"%"'
    
    auto_adjust_columns(ws, 3)


def generate_excel_report(analytics: AgendorAnalytics) -> BytesIO:
    """
    Gera relatório Excel completo com múltiplas abas
    
    Returns:
        BytesIO com o arquivo Excel
    """
    wb = Workbook()
    
    # Remover aba padrão
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Criar aba de capa
    ws_capa = wb.create_sheet("📄 Capa", 0)
    ws_capa['A1'] = "RELATÓRIO GERENCIAL"
    ws_capa['A1'].font = Font(bold=True, size=20)
    ws_capa['A3'] = f"Data de Geração: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_capa['A4'] = f"Total de Negócios: {len(analytics.df_deals)}"
    ws_capa['A6'] = "Este relatório contém análises detalhadas de:"
    ws_capa['A7'] = "• KPIs Principais"
    ws_capa['A8'] = "• Performance por Vendedor"
    ws_capa['A9'] = "• Top Clientes"
    ws_capa['A10'] = "• Funil de Conversão"
    
    # Criar abas
    create_kpi_sheet(wb, analytics)
    create_sellers_sheet(wb, analytics)
    create_top_customers_sheet(wb, analytics)
    create_funnel_sheet(wb, analytics)
    
    # Salvar em BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
